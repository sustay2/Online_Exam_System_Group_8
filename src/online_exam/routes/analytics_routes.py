from datetime import datetime
from io import BytesIO

from flask import Blueprint, render_template, send_file

from online_exam.models.exam import Exam
from online_exam.models.submission import Submission

analytics_bp = Blueprint("analytics", __name__, url_prefix="/analytics")


@analytics_bp.route("/exams/<int:exam_id>/report")
def exam_report(exam_id):
    """Display performance analytics report for an exam."""
    exam = Exam.query.get_or_404(exam_id)

    # Get all submissions for this exam (most recent first)
    submissions = (
        Submission.query.filter_by(exam_id=exam_id).order_by(Submission.submitted_at.desc()).all()
    )

    if not submissions:
        # Empty state - no submissions yet
        return render_template(
            "analytics/exam_report.html",
            exam=exam,
            submissions=[],
            total_submissions=0,
            avg_score=0,
            highest_score=0,
            lowest_score=0,
            passed=0,
            failed=0,
            pass_rate=0,
            fail_rate=0,
            score_ranges={},
        )

    # Calculate statistics
    total_submissions = len(submissions)
    total_score = sum(s.total_score for s in submissions)
    avg_score = total_score / total_submissions if total_submissions > 0 else 0
    highest_score = max(s.total_score for s in submissions)
    lowest_score = min(s.total_score for s in submissions)

    # Pass/Fail analysis (threshold: 50%)
    passed = sum(1 for s in submissions if s.percentage >= 50)
    failed = total_submissions - passed
    pass_rate = (passed / total_submissions * 100) if total_submissions > 0 else 0
    fail_rate = 100 - pass_rate

    # Score distribution (6 ranges) - use descriptive label for below 50
    score_ranges = {
        "90-100": sum(1 for s in submissions if s.percentage >= 90),
        "80-89": sum(1 for s in submissions if 80 <= s.percentage < 90),
        "70-79": sum(1 for s in submissions if 70 <= s.percentage < 80),
        "60-69": sum(1 for s in submissions if 60 <= s.percentage < 70),
        "50-59": sum(1 for s in submissions if 50 <= s.percentage < 60),
        "Below 50": sum(1 for s in submissions if s.percentage < 50),
    }

    return render_template(
        "analytics/exam_report.html",
        exam=exam,
        submissions=submissions,
        total_submissions=total_submissions,
        avg_score=avg_score,
        highest_score=highest_score,
        lowest_score=lowest_score,
        passed=passed,
        failed=failed,
        pass_rate=pass_rate,
        fail_rate=fail_rate,
        score_ranges=score_ranges,
    )


@analytics_bp.route("/exams/<int:exam_id>/export")
def export_exam_results(exam_id):
    """Export exam results to Excel (.xlsx) file."""
    exam = Exam.query.get_or_404(exam_id)
    submissions = (
        Submission.query.filter_by(exam_id=exam_id).order_by(Submission.submitted_at.desc()).all()
    )

    # Check if openpyxl is available
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        return (
            "openpyxl is not installed. Install it with: pip install openpyxl",
            500,
        )

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Exam Results"

    # Header section (rows 1-4)
    ws.merge_cells("A1:G1")
    ws["A1"] = exam.title
    ws["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # Exam info
    ws["A2"] = "Description:"
    ws["A2"].font = Font(bold=True)
    ws["B2"] = exam.description or "N/A"

    ws["A3"] = "Total Submissions:"
    ws["A3"].font = Font(bold=True)
    ws["B3"] = len(submissions)

    ws["A4"] = "Generated:"
    ws["A4"].font = Font(bold=True)
    ws["B4"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # Statistics
    if submissions:
        avg_score = sum(s.total_score for s in submissions) / len(submissions)
        high_score = max(s.total_score for s in submissions)
        low_score = min(s.total_score for s in submissions)
        pass_rate = sum(1 for s in submissions if s.percentage >= 50) / len(submissions) * 100

        ws["D2"] = "Average Score:"
        ws["D2"].font = Font(bold=True)
        ws["E2"] = f"{avg_score:.2f}"

        ws["D3"] = "Highest Score:"
        ws["D3"].font = Font(bold=True)
        ws["E3"] = high_score

        ws["D4"] = "Lowest Score:"
        ws["D4"].font = Font(bold=True)
        ws["E4"] = low_score

        ws["F3"] = "Pass Rate:"
        ws["F3"].font = Font(bold=True)
        ws["G3"] = f"{pass_rate:.1f}%"

    # Table headers (row 7)
    headers = ["#", "Student Name", "Student Email", "Score", "Percentage", "Status", "Submitted"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=7, column=col)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Data rows (row 8+)
    for idx, submission in enumerate(submissions, start=1):
        row = 7 + idx

        ws.cell(row=row, column=1, value=idx)
        ws.cell(row=row, column=2, value=submission.student_name)
        ws.cell(row=row, column=3, value=submission.student_name)  # Using name as email
        ws.cell(row=row, column=4, value=submission.total_score)

        # Percentage with color coding
        percentage_cell = ws.cell(row=row, column=5, value=f"{submission.percentage:.2f}%")
        if submission.percentage >= 80:
            percentage_cell.fill = PatternFill(
                start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
            )
            percentage_cell.font = Font(color="006100", bold=True)
        elif submission.percentage >= 50:
            percentage_cell.fill = PatternFill(
                start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"
            )
            percentage_cell.font = Font(color="9C6500")
        else:
            percentage_cell.fill = PatternFill(
                start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
            )
            percentage_cell.font = Font(color="9C0006", bold=True)

        # Status with color coding
        status = "PASS" if submission.percentage >= 50 else "FAIL"
        status_cell = ws.cell(row=row, column=6, value=status)
        status_cell.alignment = Alignment(horizontal="center")
        status_cell.font = Font(bold=True)
        if status == "PASS":
            status_cell.fill = PatternFill(
                start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
            )
            status_cell.font = Font(color="006100", bold=True)
        else:
            status_cell.fill = PatternFill(
                start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
            )
            status_cell.font = Font(color="9C0006", bold=True)

        # Submitted date
        ws.cell(row=row, column=7, value=submission.submitted_at.strftime("%Y-%m-%d %H:%M:%S"))

    # Adjust column widths
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 15
    ws.column_dimensions["G"].width = 25

    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Generate filename
    filename = (
        f"{exam.title.replace(' ', '_')}_Results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    )

    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )
